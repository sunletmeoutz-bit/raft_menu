# -*- coding: utf-8 -*-
# app.py — Калькулятор меню сплава (мобильный компактный вид + план по дням)

import io
import os
import pandas as pd
import streamlit as st

# ================= НАСТРОЙКИ СТРАНИЦЫ И СТИЛЬ =================
st.set_page_config(
    page_title="Калькулятор меню сплава",
    page_icon="🍲",
    layout="centered",
)

st.markdown("""
<style>
/* аккуратные карточки-экспандеры */
.stExpander { border-radius: 12px; margin-bottom: 10px; }
div[data-testid="stExpander"] > details { background: #1f2937; color: #fff; border: 1px solid #374151; }
div[data-testid="stExpander"] summary { padding: 8px 12px; }
/* заголовки */
h1,h2,h3 { color: #ffcc70 !important; }
/* кнопки */
.stButton button { border-radius: 10px; background: #ff7043; color: #fff; font-weight: 600; }
.stButton button:hover { background: #ff5722; }
/* поля ввода пошире под палец */
.stNumberInput > div { min-width: 280px; }
/* таблицы */
.block-container { padding-top: 1.2rem; }
</style>
""", unsafe_allow_html=True)

# ================= ДАННЫЕ ПО УМОЛЧАНИЮ =================
@st.cache_data
def загрузить_дефолтные_блюда() -> pd.DataFrame:
    # Нормы указаны на 1 человека
    data = [
        # ───────── Завтраки (3 шт) ─────────
        ["🥣 Овсянка с сухофруктами", "Овсяные хлопья", "г", 70],
        ["🥣 Овсянка с сухофруктами", "Сухофрукты (микс)", "г", 30],
        ["🥣 Овсянка с сухофруктами", "Сахар", "г", 10],
        ["🥣 Овсянка с сухофруктами", "Соль", "г", 1],

        ["🧀 Сырники", "Творог", "г", 180],
        ["🧀 Сырники", "Яйцо куриное", "шт", 0.5],
        ["🧀 Сырники", "Мука пшеничная", "г", 25],
        ["🧀 Сырники", "Сахар", "г", 15],
        ["🧀 Сырники", "Соль", "г", 1],
        ["🧀 Сырники", "Масло растительное", "мл", 10],

        ["🍳 Омлет с овощами", "Яйцо куриное", "шт", 2],
        ["🍳 Омлет с овощами", "Молоко", "мл", 50],
        ["🍳 Омлет с овощами", "Перец сладкий", "г", 30],
        ["🍳 Омлет с овощами", "Лук репчатый", "г", 20],
        ["🍳 Омлет с овощами", "Масло растительное", "мл", 10],
        ["🍳 Омлет с овощами", "Соль", "г", 3],

        # ───────── Основные блюда ─────────
        ["🍲 Плов", "Рис", "г", 100],
        ["🍲 Плов", "Свинина/говядина", "г", 130],
        ["🍲 Плов", "Морковь", "г", 40],
        ["🍲 Плов", "Лук репчатый", "г", 25],
        ["🍲 Плов", "Масло растительное", "мл", 12],
        ["🍲 Плов", "Соль", "г", 5],
        ["🍲 Плов", "Зира/специи", "г", 1],

        ["🍖 Шулюм", "Говядина", "г", 150],
        ["🍖 Шулюм", "Картофель", "г", 150],
        ["🍖 Шулюм", "Морковь", "г", 40],
        ["🍖 Шулюм", "Лук репчатый", "г", 25],
        ["🍖 Шулюм", "Соль", "г", 6],

        ["🥣 Борщ (натуральный)", "Свекла", "г", 80],
        ["🥣 Борщ (натуральный)", "Капуста белокочанная", "г", 100],
        ["🥣 Борщ (натуральный)", "Картофель", "г", 100],
        ["🥣 Борщ (натуральный)", "Морковь", "г", 40],
        ["🥣 Борщ (натуральный)", "Лук репчатый", "г", 25],
        ["🥣 Борщ (натуральный)", "Томатная паста", "г", 20],
        ["🥣 Борщ (натуральный)", "Говядина/свинина", "г", 120],
        ["🥣 Борщ (натуральный)", "Соль", "г", 6],

        ["🍝 Паста с грибами и сливками", "Паста", "г", 100],
        ["🍝 Паста с грибами и сливками", "Грибы (шампиньоны/лесные)", "г", 120],
        ["🍝 Паста с грибами и сливками", "Сливки 20%", "мл", 80],
        ["🍝 Паста с грибами и сливками", "Лук репчатый", "г", 20],
        ["🍝 Паста с грибами и сливками", "Чеснок", "г", 5],
        ["🍝 Паста с грибами и сливками", "Масло растительное/оливковое", "мл", 10],
        ["🍝 Паста с грибами и сливками", "Соль", "г", 5],

        # ───────── Ужины (натуральные) ─────────
        ["🍗 Рис с овощами и курицей", "Рис", "г", 90],
        ["🍗 Рис с овощами и курицей", "Курица (филе)", "г", 120],
        ["🍗 Рис с овощами и курицей", "Морковь", "г", 30],
        ["🍗 Рис с овощами и курицей", "Лук репчатый", "г", 20],
        ["🍗 Рис с овощами и курицей", "Масло растительное", "мл", 10],
        ["🍗 Рис с овощами и курицей", "Соль", "г", 5],

        ["🥔 Картофель с грибами", "Картофель", "г", 200],
        ["🥔 Картофель с грибами", "Грибы (шампиньоны/лесные)", "г", 120],
        ["🥔 Картофель с грибами", "Лук репчатый", "г", 30],
        ["🥔 Картофель с грибами", "Масло растительное", "мл", 15],
        ["🥔 Картофель с грибами", "Соль", "г", 5],

        ["🎃 Пшённая каша с тыквой", "Пшено", "г", 80],
        ["🎃 Пшённая каша с тыквой", "Тыква", "г", 150],
        ["🎃 Пшённая каша с тыквой", "Масло сливочное", "г", 10],
        ["🎃 Пшённая каша с тыквой", "Сахар", "г", 10],
        ["🎃 Пшённая каша с тыквой", "Соль", "г", 2],

        # ───────── Резерв «по-походному» ─────────
        ["⛺ Гречка по-походному (с тушёнкой)", "Гречка", "г", 80],
        ["⛺ Гречка по-походному (с тушёнкой)", "Тушёнка говяжья", "г", 120],
        ["⛺ Гречка по-походному (с тушёнкой)", "Лук репчатый", "г", 20],
        ["⛺ Гречка по-походному (с тушёнкой)", "Соль", "г", 5],
        ["⛺ Гречка по-походному (с тушёнкой)", "Специи", "г", 1],

        ["⛺ Паста по-походному (с тушёнкой)", "Паста", "г", 100],
        ["⛺ Паста по-походному (с тушёнкой)", "Тушёнка говяжья", "г", 120],
        ["⛺ Паста по-походному (с тушёнкой)", "Лук репчатый", "г", 20],
        ["⛺ Паста по-походному (с тушёнкой)", "Томатная паста", "г", 40],
        ["⛺ Паста по-походному (с тушёнкой)", "Соль", "г", 5],
    ]
    return pd.DataFrame(data, columns=["Блюдо","Ингредиент","Ед.изм","Норма на человека"])

if "блюда" not in st.session_state:
    st.session_state["блюда"] = загрузить_дефолтные_блюда()

# ================= ВВОД ПАРАМЕТРОВ =================
st.title("🍲 Калькулятор меню сплава")

c1, c2 = st.columns(2)
with c1:
    дни = st.number_input("📅 Количество дней", min_value=1, max_value=30, value=3, step=1)
with c2:
    участники = st.number_input("👥 Количество участников", min_value=1, max_value=100, value=6, step=1)

# ---------- МЕНЮ И ИНГРЕДИЕНТЫ (сначала выбираем блюда) ----------
st.subheader("📋 Меню и ингредиенты")

# чистая база для показа
df_view = st.session_state["блюда"].dropna(
    subset=["Блюдо","Ингредиент","Ед.изм","Норма на человека"]
).copy()

# список всех блюд
all_dishes = sorted(df_view["Блюдо"].unique().tolist())

# выбор блюд для похода (мобильный friendly)
selected_dishes = st.multiselect(
    "Выберите блюда для похода",
    options=all_dishes,
    default=[],
    help="Сначала отметьте нужные блюда — ниже можно посмотреть состав.",
    placeholder="— выберите блюда —"
)

search = st.text_input("Поиск по ингредиенту", "")

if not selected_dishes:
    st.info("Выберите хотя бы одно блюдо выше, чтобы увидеть ингредиенты.")
else:
    for dish in selected_dishes:
        sub = df_view[df_view["Блюдо"] == dish][["Ингредиент","Ед.изм","Норма на человека"]].copy()
        if search.strip():
            sub = sub[sub["Ингредиент"].str.contains(search, case=False, na=False)]

        with st.expander(f"{dish} · {len(sub)} инг.", expanded=False):
            st.dataframe(sub.reset_index(drop=True), use_container_width=True, height=220)
            total_gram = sub.loc[sub["Ед.изм"].isin(["г","мл"]), "Норма на человека"].sum()
            st.caption(f"Суммарно на 1 чел (г/мл): {total_gram:g}" if total_gram else " ")

# ---------- ПЛАН ПО ДНЯМ И ПРИЁМАМ ПИЩИ (видит только выбранные блюда) ----------
st.subheader("🗓️ План по дням и приёмам пищи")

ПРИЕМЫ_ПИЩИ = ["Завтрак", "Обед", "Ужин"]

if selected_dishes:
    варианты_для_плана = ["— не выбрано —"] + selected_dishes
else:
    # если ничего не выбрано выше — позволим выбрать из всех
    варианты_для_плана = ["— не выбрано —"] + all_dishes

план_строки = []
for d in range(1, int(дни) + 1):
    st.markdown(f"День {d}")
    cols = st.columns(3)
    for i, прием in enumerate(ПРИЕМЫ_ПИЩИ):
        with cols[i]:
            key = f"план_{d}_{прием}"
            выбор = st.selectbox(прием, варианты_для_плана, index=0, key=key)
            план_строки.append({
                "День": d,
                "Приём пищи": прием,
                "Блюдо": None if выбор == "— не выбрано —" else выбор
            })
    st.divider()

план_df = pd.DataFrame(план_строки)

# ================= РАСЧЁТ ИТОГОВОЙ ЗАКУПКИ ПО ПЛАНУ =================
st.subheader("🛒 Итоговая закупка")

блюда_df = st.session_state["блюда"].dropna(subset=["Блюдо","Ингредиент","Ед.изм","Норма на человека"]).copy()

# сколько раз каждое блюдо встречается в плане
счётчики = (план_df.dropna(subset=["Блюдо"])
            .groupby("Блюдо").size()
            .rename("Сколько раз").reset_index())

# мержим рецепты с количеством раз
расчёт = блюда_df.merge(счётчики, on="Блюдо", how="left").fillna({"Сколько раз": 0})

# итого по каждой строке рецепта
расчёт["Итого по блюду"] = расчёт["Норма на человека"] * int(участники) * расчёт["Сколько раз"]

# группируем к закупке
итог = (расчёт.groupby(["Ингредиент","Ед.изм"], as_index=False)["Итого по блюду"]
        .sum()
        .rename(columns={"Итого по блюду": "Итого"})
        .sort_values("Ингредиент"))

# убираем нули
итог = итог[итог["Итого"] > 0].reset_index(drop=True)

st.dataframe(итог, use_container_width=True)

# ================= ВЫГРУЗКА EXCEL И PDF =================
st.subheader("📂 Скачать файлы")

def excel_полный(параметры, план_df, блюда_df, закупка_df):
    """Полный Excel: Инструкция / Параметры / План / Блюда / Закупка."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()

    # Инструкция
    ws = wb.active; ws.title = "Инструкция"
    строки = [
        "Меню сплава — как пользоваться",
        "1) Задайте 'Количество дней' и 'Количество участников'.",
        "2) В разделе 'План' выберите блюда для завтрака/обеда/ужина по дням.",
        "3) Редактируйте рецепты в разделе 'Меню и ингредиенты' при необходимости.",
        "4) Скачайте 'Итоговую закупку' в Excel/PDF.",
    ]
    for i, t in enumerate(строки, 1): ws.cell(i,1,t)
    ws.column_dimensions["A"].width = 110

    # Параметры
    ws_in = wb.create_sheet("Параметры")
    ws_in.append(["Параметр","Значение"])
    ws_in.append(["Дни", параметры["days"]])
    ws_in.append(["Участники", параметры["participants"]])
    ws_in.column_dimensions["A"].width = 18; ws_in.column_dimensions["B"].width = 14

    # План
    ws_p = wb.create_sheet("План")
    ws_p.append(["День","Приём пищи","Блюдо"])
    for r in план_df[["День","Приём пищи","Блюдо"]].itertuples(index=False): ws_p.append(list(r))
    for i,w in enumerate([8,16,40],1): ws_p.column_dimensions[get_column_letter(i)].width = w

    # Блюда
    ws_d = wb.create_sheet("Блюда")
    ws_d.append(["Блюдо","Ингредиент","Ед.изм","Норма на человека"])
    for r in блюда_df[["Блюдо","Ингредиент","Ед.изм","Норма на человека"]].itertuples(index=False): ws_d.append(list(r))
    for i,w in enumerate([30,32,10,18],1): ws_d.column_dimensions[get_column_letter(i)].width = w

    # Закупка
    ws_s = wb.create_sheet("Закупка")
    ws_s.append(["Ингредиент","Ед.изм","Итого"])
    for r in закупка_df[["Ингредиент","Ед.изм","Итого"]].itertuples(index=False): ws_s.append(list(r))
    for i,w in enumerate([32,10,14],1): ws_s.column_dimensions[get_column_letter(i)].width = w

    bio = io.BytesIO(); wb.save(bio); bio.seek(0); return bio

def pdf_закупка(закупка_df, font_path="DejaVuSans.ttf"):
    """PDF только с закупкой (кириллица через TTF)."""
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    font_name = "Cyrillic"
    if os.path.exists(font_path):
        pdfmetrics.registerFont(TTFont(font_name, font_path))
    else:
        font_name = "Helvetica"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    styles["Title"].fontName = font_name

    story = [Paragraph("Итоговая закупка для сплава", styles["Title"])]
    data = [["Ингредиент","Ед.изм","Итого"]] + закупка_df.astype(str).values.tolist()
    from reportlab.platypus import Table
    table = Table([list(закупка_df.columns)] + закупка_df.astype(str).values.tolist(), repeatRows=1)
    table.setStyle([
        ('FONTNAME', (0,0), (-1,-1), font_name),
        ('BACKGROUND', (0,0), (-1,0), colors.lightblue),
        ('TEXTCOLOR', (0,0), (-1,0), colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('ALIGN',(0,0),(-1,-1),'LEFT'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
    ])
    story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer

# Кнопка Excel
excel_bytes = excel_полный(
    параметры={"days": int(дни), "participants": int(участники)},
    план_df=план_df,
    блюда_df=блюда_df,
    закупка_df=итог
)
st.download_button(
    "📊 Скачать Excel (полный файл)",
    data=excel_bytes,
    file_name="меню_сплава.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)

# Кнопка PDF
try:
    pdf_bytes = pdf_закупка(итог)
    st.download_button(
        "📄 Скачать PDF (только закупка)",
        data=pdf_bytes,
        file_name="закупка.pdf",
        mime="application/pdf"
    )
except Exception as e:
    st.warning(f"PDF недоступен: {e}. Проверь, что рядом лежит DejaVuSans.ttf и установлен reportlab.")


